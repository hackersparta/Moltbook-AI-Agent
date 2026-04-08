"""
Instagram Auto-Poster for Render deployment.
Downloads carousel slides from Google Drive, posts to IG via Meta Graph API.

Flow:
  1. Download carousel_report.xlsx from Drive
  2. Find today's scheduled post (match Scheduled Date column)
  3. Download all slides for that post from Drive
  4. Upload slides to Meta as carousel containers
  5. Publish carousel to Instagram
  6. Update Excel status → re-upload to Drive

Environment variables required on Render:
  META_PAGE_ACCESS_TOKEN  - long-lived page access token
  META_IG_ACCOUNT_ID      - Instagram business account ID
  DRIVE_FOLDER_ID         - Google Drive folder ID with posts
  GOOGLE_CREDENTIALS_JSON - service account JSON (base64 encoded) OR
  DRIVE_REFRESH_TOKEN     - OAuth refresh token
  DRIVE_CLIENT_ID         - OAuth client ID
  DRIVE_CLIENT_SECRET     - OAuth client secret
"""

import os
import io
import json
import base64
import tempfile
import logging
from datetime import datetime, date

import requests
from google.oauth2.credentials import Credentials
from google.auth.transport.requests import Request
from googleapiclient.discovery import build
from googleapiclient.http import MediaIoBaseUpload
from openpyxl import load_workbook

logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s")
log = logging.getLogger(__name__)

# ── Config from env ──────────────────────────────────────────────
META_TOKEN = os.environ.get("META_PAGE_ACCESS_TOKEN", "")
IG_ACCOUNT_ID = os.environ.get("META_IG_ACCOUNT_ID", "17841445079545451")
DRIVE_FOLDER_ID = os.environ.get("DRIVE_FOLDER_ID", "1u5TWQ4T4IP8HdnO3hhPuyPaqftt46CF5")
META_API = "https://graph.facebook.com/v21.0"

# ── Google Drive auth ────────────────────────────────────────────
def get_drive_service():
    """Build Drive service from OAuth refresh token stored in env."""
    creds = Credentials(
        token=None,
        refresh_token=os.environ.get("DRIVE_REFRESH_TOKEN"),
        client_id=os.environ.get("DRIVE_CLIENT_ID"),
        client_secret=os.environ.get("DRIVE_CLIENT_SECRET"),
        token_uri="https://oauth2.googleapis.com/token",
    )
    creds.refresh(Request())
    return build("drive", "v3", credentials=creds)


def list_drive_folder(drive, folder_id):
    """List files in a Drive folder."""
    results = drive.files().list(
        q=f"'{folder_id}' in parents and trashed=false",
        fields="files(id, name, mimeType)",
        pageSize=100,
    ).execute()
    return results.get("files", [])


def download_drive_file(drive, file_id):
    """Download file content as bytes."""
    return drive.files().get_media(fileId=file_id).execute()


def upload_drive_file(drive, folder_id, name, content_bytes, mime_type):
    """Upload/replace a file in Drive folder."""
    # Check if file already exists
    existing = drive.files().list(
        q=f"'{folder_id}' in parents and name='{name}' and trashed=false",
        fields="files(id)",
    ).execute().get("files", [])

    media = MediaIoBaseUpload(io.BytesIO(content_bytes), mimetype=mime_type)

    if existing:
        # Update existing file
        return drive.files().update(
            fileId=existing[0]["id"],
            media_body=media,
        ).execute()
    else:
        # Create new file
        return drive.files().create(
            body={"name": name, "parents": [folder_id]},
            media_body=media,
            fields="id",
        ).execute()


# ── Excel helpers ────────────────────────────────────────────────
# ★ Wishlist sheet columns (1-indexed):
#  1=#, 2=Post ID (shortcode), 3=Date, 4=Likes, 5=Slides,
#  6=Post URL, 7=Caption,
#  8=Scheduled Date, 9=Cover Image, 10=Rendered Path,
#  11=Status, 12=Posted Date

COL_NUM = 1
COL_SHORTCODE = 2
COL_DATE = 3
COL_LIKES = 4
COL_SLIDES = 5
COL_URL = 6
COL_CAPTION = 7
COL_SCHEDULED = 8
COL_STATUS = 11
COL_POSTED = 12


def find_todays_post(wb):
    """Find the next unposted row in ★ Wishlist.
    
    Priority:
      1. Exact today match with status 'rendered' or 'pending'
      2. Oldest missed post (scheduled date <= today) with status 'rendered' or 'pending'
    This allows catching up on missed posts one-by-one.
    """
    ws = wb["★ Wishlist"]
    today = date.today()
    today_str = today.strftime("%Y-%m-%d")
    POSTABLE = {"rendered", "pending"}

    best_missed = None  # (row, shortcode, sched_str)

    for row in range(2, ws.max_row + 1):
        sched = ws.cell(row=row, column=COL_SCHEDULED).value
        status = ws.cell(row=row, column=COL_STATUS).value

        if sched is None:
            continue

        # Handle both date objects and strings
        if hasattr(sched, "strftime"):
            sched_date = sched if isinstance(sched, date) else sched.date()
            sched_str = sched_date.strftime("%Y-%m-%d")
        else:
            sched_str = str(sched).strip()
            try:
                sched_date = datetime.strptime(sched_str, "%Y-%m-%d").date()
            except ValueError:
                continue

        status_lower = str(status).strip().lower() if status else ""
        if status_lower not in POSTABLE:
            continue

        # Exact today match — use immediately
        if sched_str == today_str:
            shortcode = ws.cell(row=row, column=COL_SHORTCODE).value
            log.info(f"Found today's post: row={row}, shortcode={shortcode}")
            return row, shortcode

        # Missed post (scheduled before today) — track the oldest one
        if sched_date < today and best_missed is None:
            best_missed = (row, ws.cell(row=row, column=COL_SHORTCODE).value, sched_str)

    # No exact today match — catch up on oldest missed post
    if best_missed:
        row, shortcode, sched_str = best_missed
        log.info(f"Catching up missed post: row={row}, shortcode={shortcode}, scheduled={sched_str}")
        return row, shortcode

    log.info(f"No post scheduled for {today_str} and no missed posts to catch up")
    return None, None


def mark_posted(wb, row):
    """Update Status to 'posted' and set Posted Date."""
    ws = wb["★ Wishlist"]
    ws.cell(row=row, column=COL_STATUS).value = "posted"
    ws.cell(row=row, column=COL_POSTED).value = date.today().strftime("%Y-%m-%d")


# ── Meta Graph API (with rate-limit safety) ──────────────────────
import time

# Hard safety: cooldown + daily cap to protect Meta developer account
_last_attempt_time = None   # UTC timestamp of last posting attempt
_daily_attempts = 0         # number of posting attempts today
_daily_date = None          # which date the counter is for
COOLDOWN_SECONDS = 3600     # 1 hour between attempts
MAX_DAILY_ATTEMPTS = 3      # max 3 attempts per day — protects account

def _check_cooldown():
    """Return (allowed, message). Blocks if too soon or too many attempts today."""
    global _last_attempt_time, _daily_attempts, _daily_date
    now = datetime.utcnow()
    today = now.date()

    # Reset daily counter at midnight
    if _daily_date != today:
        _daily_date = today
        _daily_attempts = 0

    # Check daily cap
    if _daily_attempts >= MAX_DAILY_ATTEMPTS:
        msg = f"Daily cap reached ({MAX_DAILY_ATTEMPTS} attempts). No more Meta calls today. Try tomorrow."
        log.warning(msg)
        return False, msg

    # Check cooldown
    if _last_attempt_time:
        elapsed = (now - _last_attempt_time).total_seconds()
        if elapsed < COOLDOWN_SECONDS:
            remaining = int(COOLDOWN_SECONDS - elapsed)
            msg = f"Cooldown active. Wait {remaining}s before next attempt ({_daily_attempts}/{MAX_DAILY_ATTEMPTS} used today)."
            log.warning(msg)
            return False, msg

    return True, "OK"

def _record_attempt():
    """Record that a posting attempt was made."""
    global _last_attempt_time, _daily_attempts
    _last_attempt_time = datetime.utcnow()
    _daily_attempts += 1
    log.info(f"Meta API attempt #{_daily_attempts}/{MAX_DAILY_ATTEMPTS} today")

def _safe_meta_call(method, url, data=None, max_retries=1):
    """Make a Meta API call with rate-limit awareness. Abort on auth errors. Max 1 retry."""
    for attempt in range(max_retries + 1):
        resp = requests.request(method, url, data=data)

        # Success
        if resp.status_code == 200:
            return resp

        error_data = resp.json().get("error", {})
        code = error_data.get("code", 0)
        msg = error_data.get("message", "")
        log.warning(f"API error: HTTP {resp.status_code}, code={code}, msg={msg}")

        # Rate limit — ABORT immediately, don't waste more calls
        if code in (4, 32) or resp.status_code == 429:
            log.error(f"RATE LIMITED — aborting to protect account. msg={msg}")
            raise Exception(f"Rate limited by Meta. Stopping to protect account.")

        # Auth errors — never retry, token is bad
        if code in (190, 10, 100):
            log.error(f"AUTH ERROR (no retry): {code} - {msg}")
            raise Exception(f"Meta auth error: {msg}")

        # Other errors — one short retry
        if attempt < max_retries:
            wait = 10
            log.warning(f"API error {resp.status_code}: {msg}. Retry in {wait}s...")
            time.sleep(wait)
            continue

        # All retries exhausted
        log.error(f"FAILED after {max_retries+1} attempts: {resp.status_code} {resp.text}")
        raise Exception(f"Meta API failed: {resp.status_code} - {msg}")

    return resp


def upload_image_to_meta(image_url, caption=None, is_carousel_item=True):
    """Upload a single image as a carousel item container."""
    payload = {
        "image_url": image_url,
        "access_token": META_TOKEN,
    }
    if is_carousel_item:
        payload["is_carousel_item"] = "true"
    if caption:
        payload["caption"] = caption

    resp = _safe_meta_call("POST", f"{META_API}/{IG_ACCOUNT_ID}/media", data=payload)
    return resp.json()["id"]


def create_carousel(container_ids, caption):
    """Create carousel container from item IDs."""
    payload = {
        "media_type": "CAROUSEL",
        "children": ",".join(container_ids),
        "caption": caption,
        "access_token": META_TOKEN,
    }
    resp = _safe_meta_call("POST", f"{META_API}/{IG_ACCOUNT_ID}/media", data=payload)
    return resp.json()["id"]


def publish_container(container_id):
    """Publish a media container. Waits for processing, then publishes with retry."""
    # Wait for Meta to process all images
    time.sleep(30)
    resp = _safe_meta_call(
        "POST",
        f"{META_API}/{IG_ACCOUNT_ID}/media_publish",
        data={"creation_id": container_id, "access_token": META_TOKEN},
        max_retries=1,
    )
    return resp.json()["id"]


def get_public_url(drive, file_id):
    """Make file publicly readable and return direct image URL at FULL resolution."""
    try:
        drive.permissions().create(
            fileId=file_id,
            body={"type": "anyone", "role": "reader"},
        ).execute()
    except Exception:
        pass  # May already be shared
    # =s0 forces original resolution (no Google compression/resize)
    return f"https://lh3.googleusercontent.com/d/{file_id}=s0"


def _already_posted_on_ig(caption_snippet):
    """Check IG recent media to see if a post with this caption already exists.
    Prevents duplicates when Meta returns error but actually publishes."""
    if not caption_snippet:
        return False
    try:
        # Get last 10 posts from IG
        resp = requests.get(
            f"{META_API}/{IG_ACCOUNT_ID}/media",
            params={
                "fields": "caption,timestamp",
                "limit": 10,
                "access_token": META_TOKEN,
            },
        )
        if resp.status_code != 200:
            log.warning(f"Could not check recent posts: {resp.status_code}")
            return False
        posts = resp.json().get("data", [])
        # Check first 80 chars of caption (enough to match uniquely)
        snippet = caption_snippet[:80].strip()
        for p in posts:
            if p.get("caption", "")[:80].strip() == snippet:
                log.warning(f"DUPLICATE DETECTED — this caption already posted on IG")
                return True
        return False
    except Exception as e:
        log.warning(f"Duplicate check failed: {e}")
        return False


# ── Main pipeline ────────────────────────────────────────────────
def run_daily_post():
    """Main function: check today's schedule, download slides, post to IG."""
    log.info("=== IG Auto-Poster starting ===")

    # Hard safety check — cooldown + daily cap
    allowed, reason = _check_cooldown()
    if not allowed:
        return {"status": "blocked", "message": reason}

    if not META_TOKEN:
        log.error("META_PAGE_ACCESS_TOKEN not set")
        return {"status": "error", "message": "META_PAGE_ACCESS_TOKEN not set"}

    drive = get_drive_service()

    # 1. Find and download carousel_report.xlsx
    log.info("Downloading carousel_report.xlsx from Drive...")
    root_files = list_drive_folder(drive, DRIVE_FOLDER_ID)
    excel_file = next((f for f in root_files if f["name"] == "carousel_report.xlsx"), None)
    if not excel_file:
        log.error("carousel_report.xlsx not found in Drive")
        return {"status": "error", "message": "carousel_report.xlsx not found"}

    excel_bytes = download_drive_file(drive, excel_file["id"])
    wb = load_workbook(io.BytesIO(excel_bytes))

    # 2. Find today's scheduled post
    row, shortcode = find_todays_post(wb)
    if row is None:
        return {"status": "skipped", "message": "No post scheduled for today"}

    # 3. Find post folder on Drive
    post_folder = next((f for f in root_files if f["name"] == shortcode), None)
    if not post_folder:
        log.error(f"Folder '{shortcode}' not found on Drive")
        return {"status": "error", "message": f"Folder '{shortcode}' not found"}

    # 4. List slides and caption
    post_files = list_drive_folder(drive, post_folder["id"])
    slides = sorted([f for f in post_files if f["name"].startswith("slide_") and f["name"].endswith(".png")],
                     key=lambda x: x["name"])
    caption_file = next((f for f in post_files if f["name"] == "caption.txt"), None)

    if not slides:
        log.error(f"No slides found in {shortcode}")
        return {"status": "error", "message": f"No slides in {shortcode}"}

    caption = ""
    if caption_file:
        caption = download_drive_file(drive, caption_file["id"]).decode("utf-8")

    log.info(f"Posting {shortcode}: {len(slides)} slides, caption={len(caption)} chars")

    # DUPLICATE CHECK — prevents reposting if Meta published but returned error
    if _already_posted_on_ig(caption):
        log.warning(f"Post {shortcode} already exists on IG! Marking as posted.")
        mark_posted(wb, row)
        buf = io.BytesIO()
        wb.save(buf)
        upload_drive_file(
            drive, DRIVE_FOLDER_ID, "carousel_report.xlsx",
            buf.getvalue(),
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        return {"status": "skipped", "message": f"{shortcode} already posted on IG (duplicate prevented)"}

    # 5. Make slides public and upload to Meta (with 2s gap between uploads)
    _record_attempt()  # Count this as an attempt BEFORE touching Meta API
    try:
        container_ids = []
        for slide in slides:
            url = get_public_url(drive, slide["id"])
            log.info(f"  Uploading {slide['name']}...")
            cid = upload_image_to_meta(url)
            container_ids.append(cid)
            log.info(f"  Container: {cid}")
            time.sleep(2)  # Gentle gap to avoid rate limits

        # 6. Create carousel and publish
        carousel_id = create_carousel(container_ids, caption)
        log.info(f"Carousel container: {carousel_id}")

        media_id = publish_container(carousel_id)
        log.info(f"Published! Media ID: {media_id}")

    except Exception as e:
        # Mark as "error" so we don't retry automatically and spam Meta
        ws = wb["★ Wishlist"]
        ws.cell(row=row, column=COL_STATUS).value = "error"
        buf = io.BytesIO()
        wb.save(buf)
        upload_drive_file(
            drive, DRIVE_FOLDER_ID, "carousel_report.xlsx",
            buf.getvalue(),
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        log.error(f"POSTING FAILED for {shortcode}: {e}")
        return {"status": "error", "shortcode": shortcode, "message": str(e)}

    # 7. Update Excel and re-upload
    mark_posted(wb, row)
    buf = io.BytesIO()
    wb.save(buf)
    upload_drive_file(
        drive, DRIVE_FOLDER_ID, "carousel_report.xlsx",
        buf.getvalue(),
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    log.info("Excel updated on Drive")

    return {
        "status": "posted",
        "shortcode": shortcode,
        "slides": len(slides),
        "media_id": media_id,
    }


if __name__ == "__main__":
    result = run_daily_post()
    print(json.dumps(result, indent=2))
