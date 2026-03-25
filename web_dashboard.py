from flask import Flask, jsonify, render_template, request, Response
from flask_cors import CORS
import json
from datetime import datetime
import os

app = Flask(__name__)

import sys
sys.path.append(os.path.dirname(__file__))
from analyze_ideas import analyze_feasibility, YOUR_SKILLS

def load_knowledge():
    try:
        with open('knowledge.json', 'r') as f:
            return json.load(f)
    except FileNotFoundError:
        return {"posts": [], "metadata": {}}

def load_seen():
    try:
        with open('seen_posts.json', 'r') as f:
            return json.load(f)
    except FileNotFoundError:
        return {"seen": [], "last_check": None}

def load_agent_state():
    try:
        with open('agent_state.json', 'r') as f:
            return json.load(f)
    except (FileNotFoundError, json.JSONDecodeError):
        return {}

@app.route('/')
def index():
    return render_template('dashboard.html')

@app.route('/api/stats')
def get_stats():
    """Get overall statistics"""
    knowledge = load_knowledge()
    seen = load_seen()
    
    posts = knowledge.get('posts', [])
    seen_ids = set(seen.get('seen', []))
    new_count = len([p for p in posts if p.get('id') not in seen_ids])
    
    return jsonify({
        'total_posts': len(posts),
        'new_posts': new_count,
        'seen_posts': len(seen_ids),
        'last_check': seen.get('last_check'),
        'last_updated': knowledge.get('metadata', {}).get('last_updated')
    })

@app.route('/api/ideas')
def get_ideas():
    """Get all ideas with AI analysis"""
    
    # Try to load AI-analyzed ideas first
    try:
        with open('ai_analyzed_ideas.json', 'r') as f:
            analyzed = json.load(f)
        
        # Transform to match frontend format
        ideas = []
        for item in analyzed:
            post = item['post']
            ai = item['ai_analysis']
            
            ideas.append({
                'id': post.get('id'),
                'title': post.get('title', 'Untitled'),
                'author': post.get('author', 'Unknown'),
                'submolt': post.get('submolt', 'general'),
                'content': post.get('content', ''),
                'url': post.get('url', ''),
                'score': post.get('score', 0),
                'comment_count': post.get('comment_count', 0),
                'created_at': post.get('created_at', ''),
                'saved_at': post.get('saved_at', ''),
                'analysis': {
                    'feasibility_score': ai.get('feasibility_score', 0),
                    'verdict': get_verdict(ai.get('feasibility_score', 0)),
                    'recommendation': ai.get('recommendation', ''),
                    'reasons_can_do': ai.get('skill_match_details', []),
                    'reasons_cant_do': ai.get('missing_skills', []),
                    'time_estimate': ai.get('time_to_build', 'Unknown'),
                    'difficulty': ai.get('difficulty', 'Unknown'),
                    'profit_potential': ai.get('profit_potential', 'Unknown'),
                    'is_ai_analysis': True,  # Flag to show this is AI-powered
                    'business_model': ai.get('business_model', 'Unknown'),
                    'tags': ai.get('tags', []),
                    'key_insights': ai.get('key_insights', [])
                }
            })
        
        return jsonify(ideas)
        
    except FileNotFoundError:
        # Fall back to knowledge base with keyword analysis
        knowledge = load_knowledge()
        posts = knowledge.get('posts', [])
        
        analyzed = []
        for post in posts:
            analysis = analyze_feasibility(post)
            analyzed.append({
                'id': post.get('id'),
                'title': post.get('title', 'Untitled'),
                'author': post.get('author', 'Unknown'),
                'submolt': post.get('submolt', 'general'),
                'content': post.get('content', ''),
                'url': post.get('url', ''),
                'score': post.get('score', 0),
                'comment_count': post.get('comment_count', 0),
                'created_at': post.get('created_at', ''),
                'saved_at': post.get('saved_at', ''),
                'analysis': {
                    'feasibility_score': analysis['score'],
                    'verdict': analysis['verdict'],
                    'recommendation': analysis['recommendation'],
                    'reasons_can_do': analysis['reasons_can_do'],
                    'reasons_cant_do': analysis['reasons_cant_do'],
                    'time_estimate': analysis['time_estimate'],
                    'difficulty': analysis['difficulty'],
                    'profit_potential': analysis['profit_potential'],
                    'is_ai_analysis': False  # Flag to show this is keyword-based
                }
            })
        
        analyzed.sort(key=lambda x: x['analysis']['feasibility_score'], reverse=True)
        return jsonify(analyzed)

def get_verdict(score):
    """Get verdict based on score"""
    if score >= 70:
        return "🟢 HIGHLY FEASIBLE"
    elif score >= 50:
        return "🟢 FEASIBLE"
    elif score >= 30:
        return "🟡 FEASIBLE - Consider Building"
    elif score >= 10:
        return "🟠 CHALLENGING"
    else:
        return "🔴 NOT RECOMMENDED"

@app.route('/api/new')
def get_new_ideas():
    """Get only new ideas"""
    knowledge = load_knowledge()
    seen = load_seen()
    
    posts = knowledge.get('posts', [])
    seen_ids = set(seen.get('seen', []))
    
    new_posts = [p for p in posts if p.get('id') not in seen_ids]
    
    # Analyze new posts
    analyzed = []
    for post in new_posts:
        analysis = analyze_feasibility(post)
        analyzed.append({
            'id': post.get('id'),
            'title': post.get('title', 'Untitled'),
            'author': post.get('author', 'Unknown'),
            'content': post.get('content', ''),
            'url': post.get('url', ''),
            'analysis': analysis
        })
    
    analyzed.sort(key=lambda x: x['analysis']['score'], reverse=True)
    
    return jsonify(analyzed)

@app.route('/api/my-posts')
def get_my_posts():
    """Get user's posts with comment counts"""
    try:
        with open('my_posts_with_comments.json', 'r') as f:
            data = json.load(f)
        
        posts = []
        for post in data.get('posts', []):
            posts.append({
                'id': post.get('id'),
                'content': post.get('content', ''),
                'created_at': post.get('created_at', ''),
                'score': post.get('score', 0),
                'comment_count': post.get('comment_count', 0),
                'url': post.get('url', '')
            })
        
        return jsonify({
            'total_posts': data.get('total_posts', 0),
            'total_comments': data.get('total_comments', 0),
            'collected_at': data.get('collected_at', ''),
            'posts': posts
        })
    except FileNotFoundError:
        return jsonify({
            'total_posts': 0,
            'total_comments': 0,
            'posts': []
        })

@app.route('/api/comment-opportunities')
def get_comment_opportunities():
    """Get AI-analyzed comment opportunities"""
    try:
        min_score = int(request.args.get('min_score', 70))
        
        if os.path.exists('comment_opportunities.json'):
            with open('comment_opportunities.json', 'r') as f:
                data = json.load(f)
            
            # Filter by minimum score
            opportunities = data.get('opportunities', [])
            filtered = [opp for opp in opportunities if opp.get('analysis', {}).get('feasibility_score', 0) >= min_score]
            
            return jsonify({
                'analyzed_at': data.get('analyzed_at'),
                'total_comments_analyzed': data.get('total_comments_analyzed'),
                'services_found': len(filtered),
                'opportunities': filtered
            })
        else:
            return jsonify({
                'services_found': 0,
                'opportunities': []
            })
    
    except Exception as e:
        print(f"Error in get_comment_opportunities: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/feed-opportunities')
def get_feed_opportunities():
    """Get opportunities from feed monitoring"""
    try:
        min_score = int(request.args.get('min_score', 70))
        
        if os.path.exists('feed_opportunities.json'):
            with open('feed_opportunities.json', 'r') as f:
                data = json.load(f)
            
            # Filter by minimum score
            opportunities = data.get('opportunities', [])
            filtered = [opp for opp in opportunities if opp.get('score', 0) >= min_score]
            
            return jsonify({
                'total': len(filtered),
                'scanned_at': data.get('scanned_at'),
                'opportunities': filtered
            })
        else:
            return jsonify({'total': 0, 'opportunities': []})
    
    except Exception as e:
        print(f"Error in get_feed_opportunities: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/comments')
def comments_page():
    """Comments analysis page"""
    return render_template('comments.html')

@app.route('/api/mark_seen', methods=['POST'])
def mark_seen():
    knowledge = load_knowledge()
    seen = load_seen()
    all_ids = [p.get('id') for p in knowledge.get('posts', [])]
    seen['seen'] = list(set(seen.get('seen', []) + all_ids))
    seen['last_check'] = datetime.now().isoformat()
    with open('seen_posts.json', 'w') as f:
        json.dump(seen, f, indent=2)
    return jsonify({"success": True, "marked": len(all_ids)})

# ── Agent Health & Intelligence endpoints ─────────────────────

@app.route('/api/agent-health')
def agent_health():
    """Real-time agent status: circuit breaker, daily caps, errors."""
    state = load_agent_state()
    return jsonify({
        'circuit_status': 'OPEN' if state.get('circuit_open_until') else 'OK',
        'circuit_open_until': state.get('circuit_open_until'),
        'consecutive_errors': state.get('consecutive_errors', 0),
        'upvotes_today': state.get('upvotes_today', 0),
        'comments_today': state.get('comments_today', 0),
        'api_calls_today': state.get('api_calls_today', 0),
        'api_calls_this_hour': state.get('api_calls_this_hour', 0),
        'counter_date': state.get('counter_date', ''),
    })

@app.route('/api/digest')
def get_digest():
    """Latest daily intelligence digest."""
    state = load_agent_state()
    digest = state.get('daily_digest', {})
    if not digest:
        # Build one on the fly from knowledge base
        knowledge = load_knowledge()
        today = datetime.now().strftime('%Y-%m-%d')
        todays = [p for p in knowledge.get('posts', []) if p.get('saved_at', '').startswith(today)]
        todays.sort(key=lambda x: x.get('intelligence_score', 0), reverse=True)
        digest = {
            'date': today,
            'ideas_saved': len(todays),
            'top_ideas': todays[:5],
            'trends': [],
        }
    return jsonify(digest)

@app.route('/api/trends')
def get_trends():
    """Keyword trend data for charting."""
    state = load_agent_state()
    trend_history = state.get('trend_history', {})

    # Return last 14 days
    from datetime import timedelta
    today = datetime.now()
    data = {}
    for i in range(14):
        day = (today - timedelta(days=i)).strftime('%Y-%m-%d')
        data[day] = trend_history.get(day, {})
    return jsonify(data)

@app.route('/api/top-ideas')
def get_top_ideas():
    """Knowledge base sorted by intelligence_score."""
    knowledge = load_knowledge()
    posts = knowledge.get('posts', [])
    # Sort by intelligence_score (new field), fallback to 0
    posts.sort(key=lambda x: x.get('intelligence_score', 0), reverse=True)
    return jsonify(posts[:20])

# ── Instagram Auto-Poster routes ─────────────────────────────────

@app.route('/ig')
@app.route('/ig/')
def ig_dashboard():
    """IG Auto-Poster dashboard page — inline HTML to avoid template path issues."""
    html = '''<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8"><meta name="viewport" content="width=device-width,initial-scale=1.0">
<title>IG Auto-Poster | @thriveaiwithnirmal</title>
<style>
*{margin:0;padding:0;box-sizing:border-box}
body{font-family:-apple-system,BlinkMacSystemFont,'Segoe UI',Roboto,sans-serif;background:#0a0a0a;color:#e0e0e0;min-height:100vh}
.header{background:linear-gradient(135deg,#1a1a2e,#16213e);border-bottom:2px solid #F5A623;padding:24px 32px;display:flex;justify-content:space-between;align-items:center}
.header h1{font-size:24px;color:#F5A623}
.header .brand{color:#888;font-size:14px}
.stats-bar{display:flex;gap:24px;padding:20px 32px;background:#111}
.stat-card{background:#1a1a2e;border-radius:12px;padding:16px 24px;flex:1;text-align:center;border:1px solid #222}
.stat-card .num{font-size:32px;font-weight:700;color:#F5A623}
.stat-card .label{font-size:12px;color:#888;text-transform:uppercase;margin-top:4px}
.actions{padding:16px 32px;display:flex;gap:12px}
.btn{padding:10px 24px;border:none;border-radius:8px;font-size:14px;font-weight:600;cursor:pointer;transition:all 0.2s}
.btn-post{background:#F5A623;color:#000}
.btn-post:hover{background:#e6951a}
.btn-refresh{background:#222;color:#e0e0e0;border:1px solid #444}
.btn-refresh:hover{background:#333}
.btn:disabled{opacity:0.5;cursor:not-allowed}
.toast{display:none;position:fixed;top:20px;right:20px;padding:12px 24px;border-radius:8px;font-size:14px;font-weight:600;z-index:1000;animation:slideIn 0.3s ease}
.toast.success{background:#1a5c2a;color:#4ade80;display:block}
.toast.error{background:#5c1a1a;color:#f87171;display:block}
.toast.info{background:#1a3a5c;color:#60a5fa;display:block}
@keyframes slideIn{from{transform:translateX(100%);opacity:0}to{transform:translateX(0);opacity:1}}
.table-wrap{padding:0 32px 32px}
table{width:100%;border-collapse:collapse;background:#111;border-radius:12px;overflow:hidden}
th{background:#1a1a2e;padding:12px 16px;text-align:left;font-size:12px;text-transform:uppercase;color:#888;border-bottom:1px solid #222}
td{padding:12px 16px;border-bottom:1px solid #1a1a1a;font-size:14px}
tr:hover td{background:#151520}
.badge{display:inline-block;padding:4px 12px;border-radius:20px;font-size:12px;font-weight:600}
.badge-rendered{background:#1a3a5c;color:#60a5fa}
.badge-posted{background:#1a5c2a;color:#4ade80}
.badge-pending{background:#3a3a1a;color:#fbbf24}
.shortcode{font-family:'SF Mono','Fira Code',monospace;color:#F5A623}
.loading{text-align:center;padding:60px;color:#666;font-size:16px}
.spinner{display:inline-block;width:20px;height:20px;border:2px solid #444;border-top-color:#F5A623;border-radius:50%;animation:spin 0.8s linear infinite;margin-right:8px;vertical-align:middle}
@keyframes spin{to{transform:rotate(360deg)}}
</style>
</head>
<body>
<div class="header">
<div><h1>&#x1F4F8; IG Auto-Poster</h1><div class="brand">@thriveaiwithnirmal &middot; Carousel Pipeline</div></div>
<div style="color:#666;font-size:13px" id="lastUpdated"></div>
</div>
<div class="stats-bar">
<div class="stat-card"><div class="num" id="totalPosts">-</div><div class="label">Total Posts</div></div>
<div class="stat-card"><div class="num" id="postedCount">-</div><div class="label">Posted</div></div>
<div class="stat-card"><div class="num" id="renderedCount">-</div><div class="label">Ready</div></div>
<div class="stat-card"><div class="num" id="pendingCount">-</div><div class="label">Pending</div></div>
</div>
<div class="actions">
<button class="btn btn-post" id="btnPost" onclick="triggerPost()">&#x1F680; Post Now (Today&#39;s Scheduled)</button>
<button class="btn btn-refresh" onclick="loadData()">&#x1F504; Refresh</button>
</div>
<div id="toast" class="toast"></div>
<div class="table-wrap"><div id="tableArea" class="loading"><span class="spinner"></span> Loading from Google Drive...</div></div>
<script>
function showToast(m,t){var e=document.getElementById("toast");e.className="toast "+t;e.textContent=m;setTimeout(function(){e.className="toast"},4000)}
function badgeClass(s){if(!s)return"badge-pending";s=s.toLowerCase();if(s==="posted")return"badge-posted";if(s==="rendered")return"badge-rendered";return"badge-pending"}
async function loadData(){
document.getElementById("tableArea").innerHTML='<div class="loading"><span class="spinner"></span> Loading from Google Drive...</div>';
try{
var r=await fetch("/api/ig-status");var d=await r.json();
if(d.error||d.status==="error"){document.getElementById("tableArea").innerHTML='<div class="loading" style="color:#f87171">Error: '+(d.message||d.error)+'</div>';return}
var p=d.posts||[];
var posted=p.filter(function(x){return(x.status||"").toLowerCase()==="posted"}).length;
var rendered=p.filter(function(x){return(x.status||"").toLowerCase()==="rendered"}).length;
var pending=p.filter(function(x){return !x.status||x.status.toLowerCase()==="pending"}).length;
document.getElementById("totalPosts").textContent=p.length;
document.getElementById("postedCount").textContent=posted;
document.getElementById("renderedCount").textContent=rendered;
document.getElementById("pendingCount").textContent=pending;
document.getElementById("lastUpdated").textContent="Updated: "+new Date().toLocaleTimeString();
var h='<table><thead><tr><th>#</th><th>Shortcode</th><th>Likes</th><th>Slides</th><th>Scheduled</th><th>Status</th><th>Posted</th></tr></thead><tbody>';
p.forEach(function(x){
var sc=x.scheduled&&x.scheduled!=="None"?x.scheduled:"\\u2014";
var pd=x.posted&&x.posted!=="None"?x.posted:"\\u2014";
h+="<tr><td>"+x.num+"</td><td class=\\"shortcode\\">"+x.shortcode+"</td><td>"+x.likes+"</td><td>"+x.slides+"</td><td>"+sc+"</td><td><span class=\\"badge "+badgeClass(x.status)+"\\">"+x.status+"</span></td><td>"+pd+"</td></tr>"});
h+="</tbody></table>";document.getElementById("tableArea").innerHTML=h;
}catch(e){document.getElementById("tableArea").innerHTML='<div class="loading" style="color:#f87171">Failed: '+e.message+'</div>'}}
async function triggerPost(){
var b=document.getElementById("btnPost");b.disabled=true;b.textContent="\\u23F3 Posting...";showToast("Triggering post in background...","info");
try{var r=await fetch("/api/ig-cron");var d=await r.json();
if(d.status==="busy"){showToast(d.message,"info");b.disabled=false;b.textContent="\\u1F680 Post Now";return}
if(d.status==="started"){showToast("Posting started! Checking progress...","info");pollResult(b);return}
showToast("Unexpected: "+(d.message||JSON.stringify(d)),"error");b.disabled=false;b.textContent="\\u1F680 Post Now";
}catch(e){showToast("Failed: "+e.message,"error");b.disabled=false;b.textContent="\\u1F680 Post Now"}}
function pollResult(b){
var iv=setInterval(async function(){
try{var r=await fetch("/api/ig-post-result");var d=await r.json();
if(d.running){b.textContent="\\u23F3 Posting... (working)";return}
clearInterval(iv);
var res=d.result||{};
if(res.status==="posted"){showToast("\\u2705 Posted! "+res.shortcode+" ("+res.slides+" slides)","success")}
else if(res.status==="skipped"){showToast(res.message||"Skipped","info")}
else if(res.status==="blocked"){showToast("\\u1F6D1 "+res.message,"error")}
else{showToast("Error: "+(res.message||"Unknown"),"error")}
b.disabled=false;b.textContent="\\u1F680 Post Now";loadData();
}catch(e){clearInterval(iv);showToast("Poll error: "+e.message,"error");b.disabled=false;b.textContent="\\u1F680 Post Now"}
},5000)}
loadData();
</script>
</body></html>'''
    return Response(html, mimetype='text/html')

# ── Background posting state ─────────────────────────────────────
import threading
_ig_post_state = {"running": False, "result": None, "started": None}

@app.route('/api/ig-cron')
def ig_cron():
    """Triggered by Render cron or external scheduler. Runs posting in background thread."""
    if _ig_post_state["running"]:
        return jsonify({"status": "busy", "message": "A post is already in progress. Please wait."})
    
    def _run():
        try:
            from ig_auto_poster import run_daily_post
            _ig_post_state["result"] = run_daily_post()
        except Exception as e:
            _ig_post_state["result"] = {"status": "error", "message": str(e)}
        finally:
            _ig_post_state["running"] = False

    _ig_post_state["running"] = True
    _ig_post_state["result"] = None
    _ig_post_state["started"] = datetime.utcnow().isoformat()
    threading.Thread(target=_run, daemon=True).start()
    return jsonify({"status": "started", "message": "Posting started in background. Refresh in ~60s."})

@app.route('/api/ig-post-result')
def ig_post_result():
    """Check result of background posting job."""
    return jsonify({
        "running": _ig_post_state["running"],
        "result": _ig_post_state["result"],
        "started": _ig_post_state["started"],
    })

@app.route('/api/ig-status')
def ig_status():
    """Check posting status — reads Excel from Drive."""
    try:
        from ig_auto_poster import get_drive_service, list_drive_folder, download_drive_file, DRIVE_FOLDER_ID
        import io as _io
        from openpyxl import load_workbook as _load_wb

        drive = get_drive_service()
        root_files = list_drive_folder(drive, DRIVE_FOLDER_ID)
        excel_file = next((f for f in root_files if f["name"] == "carousel_report.xlsx"), None)
        if not excel_file:
            return jsonify({"error": "Excel not found on Drive"})

        data = download_drive_file(drive, excel_file["id"])
        wb = _load_wb(_io.BytesIO(data))
        ws = wb["★ Wishlist"]

        posts = []
        for row in range(2, ws.max_row + 1):
            num = ws.cell(row=row, column=1).value
            if not num:
                continue
            posts.append({
                "num": num,
                "shortcode": ws.cell(row=row, column=2).value,
                "likes": ws.cell(row=row, column=4).value or 0,
                "slides": ws.cell(row=row, column=5).value or 0,
                "scheduled": str(ws.cell(row=row, column=8).value or ""),
                "status": ws.cell(row=row, column=11).value or "pending",
                "posted": str(ws.cell(row=row, column=12).value or ""),
            })

        return jsonify({"posts": posts, "total": len(posts)})
    except Exception as e:
        return jsonify({"status": "error", "message": str(e)}), 500

# ── SDS Auto-Poster ──────────────────────────────────────────────
_sds_post_state = {"running": False, "result": None, "started": None}

@app.route('/sds')
@app.route('/sds/')
def sds_dashboard():
    return render_template('sds_dashboard.html')

@app.route('/api/sds-cron')
def sds_cron():
    """Triggered by cron-job.org. Runs SDS posting in background thread."""
    if _sds_post_state["running"]:
        return jsonify({"status": "busy", "message": "SDS post already in progress."})

    slot = request.args.get("slot")

    def _run():
        try:
            from sds_poster import run_sds_post
            _sds_post_state["result"] = run_sds_post(slot)
        except Exception as e:
            _sds_post_state["result"] = {"status": "error", "message": str(e)}
        finally:
            _sds_post_state["running"] = False

    _sds_post_state["running"] = True
    _sds_post_state["result"] = None
    _sds_post_state["started"] = datetime.utcnow().isoformat()
    threading.Thread(target=_run, daemon=True).start()
    return jsonify({"status": "started", "message": f"SDS posting ({slot or 'auto'}) started in background."})

@app.route('/api/sds-post-result')
def sds_post_result():
    return jsonify({
        "running": _sds_post_state["running"],
        "result": _sds_post_state["result"],
        "started": _sds_post_state["started"],
    })

@app.route('/api/sds/status')
def sds_status():
    """Get all SDS posts with status for dashboard."""
    try:
        from sds_poster import get_sds_status
        return jsonify(get_sds_status())
    except Exception as e:
        return jsonify({"status": "error", "message": str(e)}), 500


# ── Learn AI Today FB Poster ──────────────────────────────────────
_lat_post_state = {"running": False, "result": None, "started": None}

@app.route('/lat')
@app.route('/lat/')
def lat_dashboard():
    """Learn AI Today FB Auto-Poster dashboard — custom design."""
    html = '''<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8"><meta name="viewport" content="width=device-width,initial-scale=1.0">
<title>Learn AI Today | FB Auto-Poster</title>
<style>
*{margin:0;padding:0;box-sizing:border-box}
body{font-family:'Segoe UI',system-ui,-apple-system,sans-serif;background:#0F172A;color:#E2E8F0;min-height:100vh}
.nav{background:#0F172A;border-bottom:1px solid #1E293B;padding:16px 32px;display:flex;justify-content:space-between;align-items:center}
.nav-brand{display:flex;align-items:center;gap:12px}
.nav-brand h1{font-size:20px;color:#38BDF8;font-weight:700}
.nav-brand span{color:#64748B;font-size:13px}
.nav-right{color:#475569;font-size:13px}
.grid{display:grid;grid-template-columns:repeat(4,1fr);gap:16px;padding:24px 32px}
.card{background:#1E293B;border-radius:12px;padding:20px;border:1px solid #334155}
.card .val{font-size:36px;font-weight:800;color:#38BDF8;margin:4px 0}
.card .lbl{font-size:11px;color:#64748B;text-transform:uppercase;letter-spacing:1px}
.card.accent{border-color:#0369A1;background:linear-gradient(135deg,#0C4A6E,#1E293B)}
.toolbar{padding:8px 32px 16px;display:flex;gap:12px;align-items:center}
.btn{padding:10px 22px;border:none;border-radius:8px;font-size:13px;font-weight:600;cursor:pointer;transition:all 0.15s}
.btn-primary{background:#0369A1;color:#fff}.btn-primary:hover{background:#0284C7}
.btn-secondary{background:#1E293B;color:#94A3B8;border:1px solid #334155}.btn-secondary:hover{background:#334155}
.btn:disabled{opacity:0.4;cursor:not-allowed}
.alert{margin:0 32px 16px;padding:12px 20px;border-radius:8px;font-size:13px;display:none}
.alert-ok{background:#064E3B;color:#6EE7B7;border:1px solid #065F46;display:block}
.alert-err{background:#7F1D1D;color:#FCA5A5;border:1px solid #991B1B;display:block}
.alert-info{background:#0C4A6E;color:#7DD3FC;border:1px solid #0369A1;display:block}
.tbl-wrap{padding:0 32px 48px}
table{width:100%;border-collapse:separate;border-spacing:0;background:#1E293B;border-radius:12px;overflow:hidden;border:1px solid #334155}
th{background:#0F172A;padding:12px 16px;text-align:left;font-size:11px;text-transform:uppercase;letter-spacing:1px;color:#64748B;border-bottom:1px solid #334155}
td{padding:12px 16px;border-bottom:1px solid #1E293B;font-size:13px;color:#CBD5E1}
tr:last-child td{border-bottom:none}
tr:hover td{background:#253449}
.pill{display:inline-block;padding:3px 12px;border-radius:20px;font-size:11px;font-weight:600}
.pill-ready{background:#164E63;color:#67E8F9}
.pill-posted{background:#064E3B;color:#6EE7B7}
.pill-pending{background:#3B2F14;color:#FCD34D}
.img-name{font-family:'Cascadia Code','Fira Code',monospace;color:#38BDF8;font-size:12px}
.caption-preview{max-width:300px;white-space:nowrap;overflow:hidden;text-overflow:ellipsis;color:#94A3B8;font-size:12px}
.load{text-align:center;padding:48px;color:#475569}
.spin{display:inline-block;width:18px;height:18px;border:2px solid #334155;border-top-color:#38BDF8;border-radius:50%;animation:sp .7s linear infinite;margin-right:8px;vertical-align:middle}
@keyframes sp{to{transform:rotate(360deg)}}
.footer{text-align:center;padding:24px;color:#334155;font-size:12px}
</style>
</head>
<body>
<div class="nav"><div class="nav-brand"><h1>&#x1F4D8; Learn AI Today</h1><span>Facebook Auto-Poster</span></div><div class="nav-right" id="ts"></div></div>
<div class="grid">
<div class="card accent"><div class="lbl">Total Posts</div><div class="val" id="total">-</div></div>
<div class="card"><div class="lbl">Posted</div><div class="val" id="posted">-</div></div>
<div class="card"><div class="lbl">Ready</div><div class="val" id="ready">-</div></div>
<div class="card"><div class="lbl">Next Post</div><div class="val" id="nextDay" style="font-size:18px">-</div></div>
</div>
<div class="toolbar">
<button class="btn btn-primary" id="btnPost" onclick="doPost()">&#x1F680; Post Today&#39;s</button>
<button class="btn btn-secondary" onclick="load()">&#x21BB; Refresh</button>
</div>
<div id="alert" class="alert"></div>
<div class="tbl-wrap"><div id="tbl" class="load"><span class="spin"></span>Loading schedule from Drive...</div></div>
<div class="footer">Learn AI Today &middot; Automated by Nirmal&#39;s AI Pipeline</div>
<script>
function msg(t,c){var a=document.getElementById("alert");a.className="alert alert-"+c;a.textContent=t;setTimeout(function(){a.className="alert"},5000)}
function pill(s){if(!s)return"pill-pending";s=s.toLowerCase();return s==="posted"?"pill-posted":s==="ready"?"pill-ready":"pill-pending"}
async function load(){
document.getElementById("tbl").innerHTML='<div class="load"><span class="spin"></span>Loading...</div>';
try{var r=await fetch("/api/lat-status");var d=await r.json();
if(d.status==="error"){document.getElementById("tbl").innerHTML='<div class="load" style="color:#FCA5A5">'+d.message+'</div>';return}
var p=d.posts||[];
var posted=p.filter(function(x){return(x.status||"").toLowerCase()==="posted"}).length;
var ready=p.filter(function(x){return(x.status||"").toLowerCase()==="ready"}).length;
var next=p.find(function(x){return(x.status||"").toLowerCase()==="ready"});
document.getElementById("total").textContent=p.length;
document.getElementById("posted").textContent=posted;
document.getElementById("ready").textContent=ready;
document.getElementById("nextDay").textContent=next?("Day "+next.day+" \\u2022 "+next.date):"All done!";
document.getElementById("ts").textContent="Updated "+new Date().toLocaleTimeString();
var h='<table><thead><tr><th>Day</th><th>Date</th><th>Title</th><th>Image</th><th>Caption</th><th>Status</th><th>Posted</th></tr></thead><tbody>';
p.forEach(function(x){
h+="<tr><td>"+x.day+"</td><td>"+x.date+"</td><td>"+x.title+"</td>";
h+='<td class="img-name">'+(x.image_file||"")+(x.image_exists?"  \\u2705":"  \\u274C")+"</td>";
h+='<td class="caption-preview">'+(x.caption||"").substring(0,60)+"...</td>";
h+='<td><span class="pill '+pill(x.status)+'">'+(x.status||"pending")+"</span></td>";
h+="<td>"+(x.posted_date||"\\u2014")+"</td></tr>"});
h+="</tbody></table>";document.getElementById("tbl").innerHTML=h;
}catch(e){document.getElementById("tbl").innerHTML='<div class="load" style="color:#FCA5A5">'+e.message+"</div>"}}
async function doPost(){
var b=document.getElementById("btnPost");b.disabled=true;b.textContent="\\u23F3 Posting...";msg("Posting to Facebook...","info");
try{var r=await fetch("/api/lat-cron");var d=await r.json();
if(d.status==="busy"){msg(d.message,"info");b.disabled=false;b.textContent="\\u1F680 Post Today's";return}
if(d.status==="started"){msg("Started! Checking...","info");poll(b);return}
msg(JSON.stringify(d),"err");b.disabled=false;b.textContent="\\u1F680 Post Today's";
}catch(e){msg(e.message,"err");b.disabled=false;b.textContent="\\u1F680 Post Today's"}}
function poll(b){var iv=setInterval(async function(){try{var r=await fetch("/api/lat-post-result");var d=await r.json();
if(d.running){return}clearInterval(iv);var res=d.result||{};
if(res.status==="posted"){msg("\\u2705 Posted Day "+res.day+" — "+res.title,"ok")}
else if(res.status==="skipped"){msg(res.message||"Skipped","info")}
else{msg(res.message||"Error","err")}
b.disabled=false;b.textContent="\\u1F680 Post Today's";load()
}catch(e){clearInterval(iv);msg(e.message,"err");b.disabled=false;b.textContent="\\u1F680 Post Today's"}},4000)}
load();
</script>
</body></html>'''
    return Response(html, mimetype='text/html')

@app.route('/api/lat-cron')
def lat_cron():
    """Triggered by cron-job.org at 11 AM IST. Runs LAT FB posting in background."""
    if _lat_post_state["running"]:
        return jsonify({"status": "busy", "message": "LAT post already in progress."})

    def _run():
        try:
            from lat_poster import run_lat_post
            _lat_post_state["result"] = run_lat_post()
        except Exception as e:
            _lat_post_state["result"] = {"status": "error", "message": str(e)}
        finally:
            _lat_post_state["running"] = False

    _lat_post_state["running"] = True
    _lat_post_state["result"] = None
    _lat_post_state["started"] = datetime.utcnow().isoformat()
    threading.Thread(target=_run, daemon=True).start()
    return jsonify({"status": "started", "message": "LAT FB posting started in background."})

@app.route('/api/lat-post-result')
def lat_post_result():
    return jsonify({
        "running": _lat_post_state["running"],
        "result": _lat_post_state["result"],
        "started": _lat_post_state["started"],
    })

@app.route('/api/lat-status')
def lat_status():
    """Get all LAT posts with status for dashboard."""
    try:
        from lat_poster import get_lat_status
        return jsonify(get_lat_status())
    except Exception as e:
        return jsonify({"status": "error", "message": str(e)}), 500


if __name__ == '__main__':
    print("=" * 70)
    print("🚀 MOLTBOOK IDEA ANALYZER - WEB DASHBOARD")
    print("=" * 70)
    print("\n📊 Starting web server...")
    print("🌐 Open your browser to: http://localhost:5000")
    print("\n💡 Press Ctrl+C to stop the server\n")
    # Run on 0.0.0.0 to be accessible from outside container
    port = int(os.environ.get('PORT', 5000))
    app.run(host='0.0.0.0', port=port, debug=True)
