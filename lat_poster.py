"""
Learn AI Today — Facebook Auto-Poster for Render deployment.
Downloads image + Excel schedule from Google Drive, posts to FB page via Meta Graph API.

Uses the SAME Meta page token as the IG poster (Learn AI Today page).
Uses a SEPARATE Drive folder for LAT-FB content.

Environment variables (Render):
  META_PAGE_ACCESS_TOKEN  - Learn AI Today page access token (shared with IG poster)
  LAT_META_PAGE_ID        - Learn AI Today Facebook page ID
  LAT_DRIVE_FOLDER_ID     - Google Drive folder with LAT-FB posts + schedule
  # Google Drive auth (shared with existing posters):
  DRIVE_REFRESH_TOKEN / DRIVE_CLIENT_ID / DRIVE_CLIENT_SECRET
"""

import os
import io
import json
import time
import logging
from datetime import datetime, date, timedelta

import requests
from google.oauth2.credentials import Credentials
from google.auth.transport.requests import Request
from googleapiclient.discovery import build
from googleapiclient.http import MediaIoBaseUpload
from openpyxl import load_workbook

logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s")
log = logging.getLogger("lat_poster")

# ── Config from env ──────────────────────────────────────────────
LAT_PAGE_TOKEN = os.environ.get("META_PAGE_ACCESS_TOKEN", "")
LAT_PAGE_ID = os.environ.get("LAT_META_PAGE_ID", "1455821964711020")
LAT_DRIVE_FOLDER_ID = os.environ.get("LAT_DRIVE_FOLDER_ID", "16ZyjL8FY_K6fosZ3lISzjKpd9gU3Tpub")
META_API = "https://graph.facebook.com/v21.0"

# Excel column mapping (1-indexed, matching lat_schedule.xlsx)
COL_DAY = 1
COL_DATE = 2
COL_TITLE = 3
COL_IMAGE = 4
COL_CAPTION = 5
COL_STATUS = 6
COL_POSTED_DATE = 7
COL_POST_ID = 8


# ── Google Drive ─────────────────────────────────────────────────
def get_drive_service():
    creds = Credentials(
        token=None,
        refresh_token=os.environ.get("DRIVE_REFRESH_TOKEN"),
        client_id=os.environ.get("DRIVE_CLIENT_ID"),
        client_secret=os.environ.get("DRIVE_CLIENT_SECRET"),
        token_uri="https://oauth2.googleapis.com/token",
    )
    creds.refresh(Request())
    return build("drive", "v3", credentials=creds)


def list_drive_folder(drive, folder_id):
    results = drive.files().list(
        q=f"'{folder_id}' in parents and trashed=false",
        fields="files(id, name, mimeType)",
        pageSize=200,
    ).execute()
    return results.get("files", [])


def download_drive_file(drive, file_id):
    return drive.files().get_media(fileId=file_id).execute()


def upload_drive_bytes(drive, folder_id, name, content_bytes, mime_type):
    existing = drive.files().list(
        q=f"'{folder_id}' in parents and name='{name}' and trashed=false",
        fields="files(id)",
    ).execute().get("files", [])
    media = MediaIoBaseUpload(io.BytesIO(content_bytes), mimetype=mime_type)
    if existing:
        return drive.files().update(fileId=existing[0]["id"], media_body=media).execute()
    return drive.files().create(
        body={"name": name, "parents": [folder_id]},
        media_body=media, fields="id",
    ).execute()


def get_public_url(drive, file_id):
    """Make file publicly readable and return a direct URL."""
    try:
        drive.permissions().create(
            fileId=file_id,
            body={"type": "anyone", "role": "reader"},
        ).execute()
    except Exception:
        pass
    return f"https://lh3.googleusercontent.com/d/{file_id}=s0"


# ── Excel helpers ────────────────────────────────────────────────
def load_schedule(drive):
    """Download and parse lat_schedule.xlsx from Drive."""
    files = list_drive_folder(drive, LAT_DRIVE_FOLDER_ID)
    excel = next((f for f in files if f["name"] == "lat_schedule.xlsx"), None)
    if not excel:
        raise FileNotFoundError("lat_schedule.xlsx not found in Drive folder")
    data = download_drive_file(drive, excel["id"])
    wb = load_workbook(io.BytesIO(data))
    return wb, excel["id"]


def get_all_posts(wb):
    """Extract all posts from the schedule."""
    ws = wb["Schedule"]
    posts = []
    for row in range(2, ws.max_row + 1):
        day = ws.cell(row, COL_DAY).value
        if day is None:
            continue
        date_val = ws.cell(row, COL_DATE).value
        if date_val and hasattr(date_val, "strftime"):
            date_str = date_val.strftime("%Y-%m-%d")
        else:
            date_str = str(date_val or "")

        posts.append({
            "row": row,
            "day": int(day),
            "date": date_str,
            "title": str(ws.cell(row, COL_TITLE).value or ""),
            "image_file": str(ws.cell(row, COL_IMAGE).value or ""),
            "caption": str(ws.cell(row, COL_CAPTION).value or ""),
            "status": str(ws.cell(row, COL_STATUS).value or "ready"),
        })
    return posts


def find_todays_post(posts):
    """Find the post scheduled for today with status='ready'."""
    today_str = date.today().strftime("%Y-%m-%d")
    for p in posts:
        if p["date"] == today_str and p["status"].lower() == "ready":
            return p
    return None


# ── Meta Graph API ───────────────────────────────────────────────
def _meta_request(method, endpoint, data=None):
    """Make request to Meta Graph API using LAT page token."""
    url = f"{META_API}{endpoint}"
    headers = {"Authorization": f"Bearer {LAT_PAGE_TOKEN}"}
    if method == "GET":
        resp = requests.get(url, headers=headers, params=data, timeout=60)
    else:
        resp = requests.post(url, headers=headers, json=data, timeout=60)
    if resp.status_code != 200:
        log.error(f"Meta API {resp.status_code}: {resp.text[:500]}")
    resp.raise_for_status()
    return resp.json()


def post_photo_to_fb(image_url, caption):
    """Upload photo to FB page and publish with caption."""
    result = _meta_request("POST", f"/{LAT_PAGE_ID}/photos", data={
        "url": image_url,
        "message": caption,
        "published": True,
    })
    return result.get("id") or result.get("post_id")


# ── Main entry point ─────────────────────────────────────────────
def run_lat_post():
    """Run LAT FB posting for today. Called by cron or manually."""
    log.info("=== LAT FB Poster starting ===")

    if not LAT_PAGE_TOKEN:
        return {"status": "error", "message": "META_PAGE_ACCESS_TOKEN not set"}

    drive = get_drive_service()
    wb, excel_id = load_schedule(drive)
    posts = get_all_posts(wb)
    post = find_todays_post(posts)

    if not post:
        log.info("No post scheduled for today (or already posted)")
        return {"status": "skipped", "message": "No ready post for today"}

    log.info(f"Today's post: Day {post['day']} — {post['title']}")

    # Find the image file on Drive
    files = list_drive_folder(drive, LAT_DRIVE_FOLDER_ID)
    image_file = next((f for f in files if f["name"] == post["image_file"]), None)
    if not image_file:
        return {"status": "error", "message": f"Image {post['image_file']} not found on Drive"}

    image_url = get_public_url(drive, image_file["id"])
    log.info(f"Image URL: {image_url}")

    # Post to FB
    post_id = post_photo_to_fb(image_url, post["caption"])
    log.info(f"FB posted! Post ID: {post_id}")

    # Update Excel
    ws = wb["Schedule"]
    ws.cell(post["row"], COL_STATUS).value = "posted"
    ws.cell(post["row"], COL_POSTED_DATE).value = date.today().strftime("%Y-%m-%d")
    ws.cell(post["row"], COL_POST_ID).value = str(post_id)

    buf = io.BytesIO()
    wb.save(buf)
    upload_drive_bytes(drive, LAT_DRIVE_FOLDER_ID, "lat_schedule.xlsx",
                       buf.getvalue(),
                       "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    log.info("Excel updated on Drive ✓")

    return {
        "status": "posted",
        "day": post["day"],
        "title": post["title"],
        "post_id": post_id,
    }


def get_lat_status():
    """Get all LAT posts status for dashboard."""
    try:
        drive = get_drive_service()
        wb, _ = load_schedule(drive)
        posts = get_all_posts(wb)

        # Check which images exist on Drive
        files = list_drive_folder(drive, LAT_DRIVE_FOLDER_ID)
        file_names = {f["name"] for f in files}

        for p in posts:
            p["image_exists"] = p["image_file"] in file_names

        return {"status": "ok", "posts": posts}
    except Exception as e:
        log.error(f"Status check failed: {e}")
        return {"status": "error", "message": str(e)}
