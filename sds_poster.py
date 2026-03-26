"""
SDS Auto-Poster for Render deployment.
Downloads images + Excel from Google Drive, posts to FB & IG via Meta Graph API.

Uses SEPARATE Meta credentials (SDS account) — does NOT touch the main Moltbook IG config.

Environment variables (Render):
  SDS_META_PAGE_TOKEN     - SDS page access token
  SDS_META_PAGE_ID        - SDS Facebook page ID
  SDS_META_IG_ID          - SDS Instagram business account ID
  SDS_DRIVE_FOLDER_ID     - Google Drive folder with SDS posts
  # Google Drive auth (shared with existing IG poster):
  DRIVE_REFRESH_TOKEN / DRIVE_CLIENT_ID / DRIVE_CLIENT_SECRET
"""

import os
import io
import json
import time
import logging
from datetime import datetime, date, timedelta

import requests
from google.oauth2.credentials import Credentials
from google.auth.transport.requests import Request
from googleapiclient.discovery import build
from googleapiclient.http import MediaIoBaseUpload
from openpyxl import load_workbook

logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s")
log = logging.getLogger("sds_poster")

# ── Config from env ──────────────────────────────────────────────
SDS_PAGE_TOKEN = os.environ.get("SDS_META_PAGE_TOKEN", "")
SDS_PAGE_ID = os.environ.get("SDS_META_PAGE_ID", "104385049240392")
SDS_IG_ID = os.environ.get("SDS_META_IG_ID", "17841400858144424")
SDS_DRIVE_FOLDER_ID = os.environ.get("SDS_DRIVE_FOLDER_ID", "1-Q-7ox49hJOJTsh8ORikKZZsAxJKjibf")
META_API = "https://graph.facebook.com/v21.0"

# Excel column mapping (1-indexed, matching content_calendar_v2.xlsx)
COL_DAY = 1
COL_DATE = 2
COL_DAY_NAME = 3
COL_TIME = 4
COL_PLATFORM = 5
COL_CONTENT_TYPE = 6
COL_TEMPLATE = 7
COL_PHOTO = 8
COL_TAMIL = 9
COL_ENGLISH = 10
COL_CAPTION = 11
COL_HASHTAGS = 12
COL_STATUS = 13


# ── Google Drive ─────────────────────────────────────────────────
def get_drive_service():
    creds = Credentials(
        token=None,
        refresh_token=os.environ.get("DRIVE_REFRESH_TOKEN"),
        client_id=os.environ.get("DRIVE_CLIENT_ID"),
        client_secret=os.environ.get("DRIVE_CLIENT_SECRET"),
        token_uri="https://oauth2.googleapis.com/token",
    )
    creds.refresh(Request())
    return build("drive", "v3", credentials=creds)


def list_drive_folder(drive, folder_id):
    results = drive.files().list(
        q=f"'{folder_id}' in parents and trashed=false",
        fields="files(id, name, mimeType)",
        pageSize=200,
    ).execute()
    return results.get("files", [])


def download_drive_file(drive, file_id):
    return drive.files().get_media(fileId=file_id).execute()


def upload_drive_bytes(drive, folder_id, name, content_bytes, mime_type):
    existing = drive.files().list(
        q=f"'{folder_id}' in parents and name='{name}' and trashed=false",
        fields="files(id)",
    ).execute().get("files", [])
    media = MediaIoBaseUpload(io.BytesIO(content_bytes), mimetype=mime_type)
    if existing:
        return drive.files().update(fileId=existing[0]["id"], media_body=media).execute()
    return drive.files().create(body={"name": name, "parents": [folder_id]}, media_body=media, fields="id").execute()


def get_public_url(drive, file_id):
    try:
        drive.permissions().create(fileId=file_id, body={"type": "anyone", "role": "reader"}).execute()
    except Exception:
        pass
    return f"https://lh3.googleusercontent.com/d/{file_id}=s0"


# ── Excel helpers ────────────────────────────────────────────────
def load_calendar(drive):
    """Download and parse the Excel calendar from Drive."""
    files = list_drive_folder(drive, SDS_DRIVE_FOLDER_ID)
    excel = next((f for f in files if f["name"] == "content_calendar_v2.xlsx"), None)
    if not excel:
        raise FileNotFoundError("content_calendar_v2.xlsx not found in Drive")
    data = download_drive_file(drive, excel["id"])
    wb = load_workbook(io.BytesIO(data))
    return wb, excel["id"]


def get_all_posts(wb):
    """Extract all posts from the calendar."""
    ws = wb.active
    posts = []
    for row in range(2, ws.max_row + 1):
        day = ws.cell(row, COL_DAY).value
        if day is None:
            continue
        date_val = ws.cell(row, COL_DATE).value
        if date_val and hasattr(date_val, "strftime"):
            date_str = date_val.strftime("%d-%b-%Y")
        else:
            date_str = str(date_val or "")

        caption = ws.cell(row, COL_CAPTION).value or ""
        hashtags = ws.cell(row, COL_HASHTAGS).value or ""
        full_caption = (caption.strip() + "\n\n" + hashtags.strip()).strip() if hashtags else caption.strip()

        posts.append({
            "row": row,
            "day": int(day),
            "date": date_str,
            "day_name": ws.cell(row, COL_DAY_NAME).value or "",
            "time": str(ws.cell(row, COL_TIME).value or ""),
            "platform": str(ws.cell(row, COL_PLATFORM).value or ""),
            "content_type": str(ws.cell(row, COL_CONTENT_TYPE).value or ""),
            "template": str(ws.cell(row, COL_TEMPLATE).value or ""),
            "tamil": str(ws.cell(row, COL_TAMIL).value or ""),
            "english": str(ws.cell(row, COL_ENGLISH).value or ""),
            "caption": full_caption,
            "status": str(ws.cell(row, COL_STATUS).value or "Pending"),
        })
    return posts


def find_posts_for_slot(posts, slot):
    """Find posts for today matching the given slot (morning/evening)."""
    today = date.today()
    today_str = today.strftime("%d-%b-%Y")

    slot_hour = 9 if slot == "morning" else 18

    matching = []
    for p in posts:
        # Parse date
        try:
            if hasattr(p["date"], "date"):
                post_date = p["date"]
            else:
                post_date = datetime.strptime(p["date"], "%d-%b-%Y").date()
        except (ValueError, AttributeError):
            continue

        if post_date != today:
            continue

        if p["status"].lower() in ("posted", "fb_posted"):
            if p["status"].lower() == "posted":
                continue

        # Match time slot
        t = p["time"].strip().upper()
        if "AM" in t or "PM" in t:
            try:
                for fmt in ("%I:%M %p", "%I:%M%p"):
                    try:
                        parsed_hour = datetime.strptime(t, fmt).hour
                        break
                    except ValueError:
                        continue
                else:
                    continue
            except Exception:
                continue
        else:
            try:
                parsed_hour = int(t.split(":")[0])
            except (ValueError, IndexError):
                continue

        if parsed_hour == slot_hour:
            matching.append(p)

    return matching


# ── Meta Graph API ───────────────────────────────────────────────
def _meta_request(method, endpoint, data=None, files=None):
    """Make request to Meta Graph API using SDS tokens."""
    url = f"{META_API}{endpoint}"
    headers = {"Authorization": f"Bearer {SDS_PAGE_TOKEN}"}
    if method == "GET":
        resp = requests.get(url, headers=headers, params=data, timeout=60)
    elif files:
        resp = requests.post(url, headers=headers, data=data, files=files, timeout=60)
    else:
        resp = requests.post(url, headers=headers, json=data, timeout=60)
    if resp.status_code != 200:
        log.error(f"Meta API {resp.status_code}: {resp.text[:300]}")
    resp.raise_for_status()
    return resp.json()


def upload_photo_to_fb(image_url):
    """Upload image to FB page (unpublished) using URL."""
    result = _meta_request("POST", f"/{SDS_PAGE_ID}/photos", data={
        "url": image_url,
        "published": "false",
        "temporary": "false",
    })
    return result["id"]


def post_to_fb(message, photo_id):
    result = _meta_request("POST", f"/{SDS_PAGE_ID}/feed", data={
        "message": message,
        "attached_media": json.dumps([{"media_fbid": photo_id}]),
    })
    return result["id"]


def create_ig_container(image_url, caption, is_carousel_item=False):
    data = {"image_url": image_url, "caption": caption}
    if is_carousel_item:
        data["is_carousel_item"] = "true"
    result = _meta_request("POST", f"/{SDS_IG_ID}/media", data=data)
    return result["id"]


def create_ig_carousel(caption, child_ids):
    result = _meta_request("POST", f"/{SDS_IG_ID}/media", data={
        "media_type": "CAROUSEL",
        "caption": caption,
        "children": ",".join(child_ids),
    })
    return result["id"]


def wait_for_container(container_id, max_wait=300):
    start = time.time()
    while time.time() - start < max_wait:
        try:
            result = _meta_request("GET", f"/{container_id}", data={"fields": "status_code"})
            if result.get("status_code") == "FINISHED":
                return True
        except Exception:
            pass
        time.sleep(5)
    return False


def publish_ig(container_id):
    result = _meta_request("POST", f"/{SDS_IG_ID}/media_publish", data={"creation_id": container_id})
    return result["id"]


# ── Posting logic ────────────────────────────────────────────────
def post_fb_and_ig(drive, day_num, caption):
    """Post single image to both FB and IG."""
    day_folder_name = f"day{day_num:02d}"
    root_files = list_drive_folder(drive, SDS_DRIVE_FOLDER_ID)
    day_folder = next((f for f in root_files if f["name"] == day_folder_name), None)
    if not day_folder:
        raise FileNotFoundError(f"Folder {day_folder_name} not found on Drive")

    files = list_drive_folder(drive, day_folder["id"])
    single_images = sorted([f for f in files if f["name"].endswith(".png") and "carousel" not in f["name"]], key=lambda x: x["name"])
    if not single_images:
        raise FileNotFoundError(f"No single image in {day_folder_name}")

    image_url = get_public_url(drive, single_images[0]["id"])
    log.info(f"Image URL: {image_url}")

    # FB
    photo_id = upload_photo_to_fb(image_url)
    fb_post_id = post_to_fb(caption, photo_id)
    log.info(f"FB posted: {fb_post_id}")

    # IG
    ig_container = create_ig_container(image_url, caption)
    if not wait_for_container(ig_container):
        log.error("IG container timeout")
        return "fb_only", fb_post_id
    ig_post_id = publish_ig(ig_container)
    log.info(f"IG posted: {ig_post_id}")

    return "both", fb_post_id


def post_ig_carousel(drive, day_num, caption):
    """Post carousel to IG only."""
    day_folder_name = f"day{day_num:02d}"
    root_files = list_drive_folder(drive, SDS_DRIVE_FOLDER_ID)
    day_folder = next((f for f in root_files if f["name"] == day_folder_name), None)
    if not day_folder:
        raise FileNotFoundError(f"Folder {day_folder_name} not found on Drive")

    files = list_drive_folder(drive, day_folder["id"])
    carousel_files = sorted([f for f in files if "carousel" in f["name"] and f["name"].endswith(".png")],
                             key=lambda x: x["name"])
    if not carousel_files:
        raise FileNotFoundError(f"No carousel images in {day_folder_name}")

    child_ids = []
    for cf in carousel_files:
        url = get_public_url(drive, cf["id"])
        cid = create_ig_container(url, caption, is_carousel_item=True)
        child_ids.append(cid)
        log.info(f"  Slide {cf['name']} -> {cid}")
        time.sleep(2)

    carousel_id = create_ig_carousel(caption, child_ids)
    if not wait_for_container(carousel_id):
        raise TimeoutError("Carousel container timeout")
    ig_post_id = publish_ig(carousel_id)
    log.info(f"IG carousel posted: {ig_post_id}")

    return "posted", ig_post_id


# ── Main entry point ─────────────────────────────────────────────
def run_sds_post(slot=None):
    """Run SDS posting for the given slot. Called by cron or manually."""
    log.info(f"=== SDS Poster starting (slot={slot}) ===")

    if not SDS_PAGE_TOKEN:
        return {"status": "error", "message": "SDS_META_PAGE_TOKEN not set"}

    drive = get_drive_service()
    wb, excel_id = load_calendar(drive)
    posts = get_all_posts(wb)

    if slot is None:
        # Determine slot from current hour (IST = UTC+5:30)
        utc_now = datetime.utcnow()
        ist_hour = (utc_now + timedelta(hours=5, minutes=30)).hour
        slot = "morning" if ist_hour < 15 else "evening"

    matching = find_posts_for_slot(posts, slot)
    if not matching:
        log.info("No matching posts for today's slot")
        return {"status": "skipped", "message": f"No pending posts for {slot} slot today"}

    results = []
    ws = wb.active

    for post in matching:
        log.info(f"Processing: Day {post['day']}, {post['platform']}, {post['time']}")
        try:
            if "FB + IG" in post["platform"]:
                fb_done = post["status"].lower() == "fb_posted"
                if fb_done:
                    # Only retry IG
                    result, pid = post_ig_carousel(drive, post["day"], post["caption"])
                else:
                    result, pid = post_fb_and_ig(drive, post["day"], post["caption"])

                if result == "both":
                    ws.cell(post["row"], COL_STATUS).value = "Posted"
                elif result == "fb_only":
                    ws.cell(post["row"], COL_STATUS).value = "FB_Posted"
                results.append({"day": post["day"], "time": post["time"], "result": result})

            elif "IG Only" in post["platform"]:
                result, pid = post_ig_carousel(drive, post["day"], post["caption"])
                ws.cell(post["row"], COL_STATUS).value = "Posted"
                results.append({"day": post["day"], "time": post["time"], "result": result})

            else:
                log.warning(f"Unknown platform: {post['platform']}")
                results.append({"day": post["day"], "time": post["time"], "result": "unknown_platform"})

        except Exception as e:
            log.error(f"Error posting day {post['day']}: {e}")
            results.append({"day": post["day"], "time": post["time"], "result": "error", "message": str(e)})

    # Save Excel back to Drive
    buf = io.BytesIO()
    wb.save(buf)
    upload_drive_bytes(drive, SDS_DRIVE_FOLDER_ID, "content_calendar_v2.xlsx",
                       buf.getvalue(),
                       "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    log.info("Excel updated on Drive")

    success_count = sum(1 for r in results if r["result"] in ("both", "posted"))
    return {
        "status": "completed",
        "slot": slot,
        "total": len(results),
        "success": success_count,
        "details": results,
    }


def get_sds_status():
    """Get all posts status for dashboard display."""
    try:
        drive = get_drive_service()
        wb, _ = load_calendar(drive)
        posts = get_all_posts(wb)

        # Get image info from Drive folders
        root_files = list_drive_folder(drive, SDS_DRIVE_FOLDER_ID)
        day_folders = {f["name"]: f["id"] for f in root_files
                       if f["mimeType"] == "application/vnd.google-apps.folder"}

        for post in posts:
            day_key = f"day{post['day']:02d}"
            folder_id = day_folders.get(day_key)
            if folder_id:
                files = list_drive_folder(drive, folder_id)
                images = [f for f in files if f["name"].endswith(".png")]
                single = [f for f in images if "carousel" not in f["name"]]
                carousel = [f for f in images if "carousel" in f["name"]]
                post["image_count"] = len(images)
                post["has_single"] = len(single) > 0
                post["has_carousel"] = len(carousel) > 0
                if single:
                    post["preview_url"] = f"https://lh3.googleusercontent.com/d/{single[0]['id']}=s400"
                elif carousel:
                    post["preview_url"] = f"https://lh3.googleusercontent.com/d/{carousel[0]['id']}=s400"
            else:
                post["image_count"] = 0
                post["has_single"] = False
                post["has_carousel"] = False

        return {"status": "ok", "posts": posts}

    except Exception as e:
        log.error(f"Error getting status: {e}")
        return {"status": "error", "message": str(e)}


if __name__ == "__main__":
    import sys
    slot = sys.argv[1] if len(sys.argv) > 1 else None
    result = run_sds_post(slot)
    print(json.dumps(result, indent=2))
